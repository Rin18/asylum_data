import openpyxl

# load excel with its path
file_name = 'admissions.xlsx'
wrkbk = openpyxl.load_workbook(file_name)
  
# get worksheet
sh = wrkbk.active
  
# iterate through data and append in category array
# if parsing through male data M = true
def parse_rows(row, sheet, category, M):
  limit = 56 if M else sh.max_row+1
  for i in range(row, limit, 14):
    for j in range(2, sheet.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        if(cell_obj.value != None):
            category.append(cell_obj.value)

# calculate percentage 
def percentage(part, whole):
    return 100 * float(part)/float(whole)


# Male data arrays
number_M = []
age_M = []
place_M = []
occupation_M = []
religion_M = []
duration_M = []
outcome_M = []
parse_rows(1, sh, number_M, True)
parse_rows(2, sh, age_M, True)
parse_rows(5, sh, place_M, True)
parse_rows(6, sh, occupation_M, True)
parse_rows(7, sh, religion_M, True)
parse_rows(12, sh, outcome_M, True)
parse_rows(14, sh, duration_M, True)

# Female data arrays
number_F = []
age_F = []
place_F = []
occupation_F = []
religion_F = []
duration_F = []
outcome_F = []
parse_rows(57, sh, number_F, False)
parse_rows(58, sh, age_F, False)
parse_rows(61, sh, place_F, False)
parse_rows(62, sh, occupation_F, False)
parse_rows(63, sh, religion_F, False)
parse_rows(68, sh, outcome_F, False)
parse_rows(70, sh, duration_F, False)


# Women stats

# Stayed over 6 months %
p_over = 0
count = 0
for i in duration_F:
    if(i > 180):
        count = count + 1
    p_over = percentage(count, len(duration_F))

# Average stay
sum_stay = 0
average_stay = 0
for i in duration_F:
    if isinstance(i, int):
        sum_stay = sum_stay + i
    average_stay = sum_stay / len(duration_F)

# Average age
sum = 0
average = 0
for i in age_F:
    if isinstance(i, int):
        sum = sum + i
    average = sum / len(age_F)

# Church %
p_CoE = 0
count = 0
for i in religion_F:
    if(i == 'CoE'):
        count = count + 1
    p_CoE = percentage(count, len(religion_F))

# Outcome %
p_recovered = 0
p_died = 0
count_d = 0
count_r = 0
for i in outcome_F:
    if(i == 'Died'):
        count_d = count_d + 1
    elif(i == 'Recovered'):
        count_r = count_r + 1
    p_died = percentage(count_d, len(outcome_F))
    p_recovered = percentage(count_r, len(outcome_F))

# Came from %
p_workhouse = 0
p_police = 0
count_w = 0
count_p = 0
for i in place_F:
    if(i == 'Workhouse'):
        count_w = count_w + 1
    elif(i == 'Brought in by police'):
        count_p = count_p + 1
    p_workhouse = percentage(count_w, len(place_F))
    p_police = percentage(count_p, len(place_F))

# Print stats women
print("Women:")
print("Percentage with length of stay over 6 months: \n", p_over)
print("Average stay: \n", average_stay)
print("Average age: \n", average)
print("Percentage CoE: \n", p_CoE)
print("Outcome: ")
print("Percentage died: ", p_died)
print("Percentage recovered: ", p_recovered)
print("Occupations: \n", set(occupation_F))
print("Came from: ")
print("Percentage workhouse: ", p_workhouse)
print("Percentage police: ", p_police)

print("\n|---------------------------------------------------------------------------------------------------------------------------------------------------------------------------|\n")


# Men stats

# Stayed over 6 months %
p_over = 0
count = 0
for i in duration_M:
    if(i > 180):
        count = count + 1
    p_over = percentage(count, len(duration_M))

# Average stay
sum_stay = 0
average_stay = 0
for i in duration_M:
    if isinstance(i, int):
        sum_stay = sum_stay + i
    average_stay = sum_stay / len(duration_M)

# Average age
sum = 0
average = 0
for i in age_M:
    if isinstance(i, int):
        sum = sum + i
    average = sum / len(age_M)

# Church %
p_CoE = 0
count = 0
for i in religion_M:
    if(i == 'CoE'):
        count = count + 1
    p_CoE = percentage(count, len(religion_M))

# Outcome %
p_recovered = 0
p_died = 0
count_d = 0
count_r = 0
for i in outcome_M:
    if(i == 'Died'):
        count_d = count_d + 1
    elif(i == 'Recovered'):
        count_r = count_r + 1
    p_died = percentage(count_d, len(outcome_M))
    p_recovered = percentage(count_r, len(outcome_M))

# Came from %
p_workhouse = 0
p_police = 0
count_w = 0
count_p = 0
for i in place_M:
    if(i == 'Workhouse'):
        count_w = count_w + 1
    elif(i == 'Brought in by police'):
        count_p = count_p + 1
    p_workhouse = percentage(count_w, len(place_M))
    p_police = percentage(count_p, len(place_M))


# Print stats men

print("Men:")
print("Percentage with length of stay over 6 months: \n", p_over)
print("Average stay: \n", average_stay)
average_stay = (sum_stay-4616) / (len(duration_M)-1)
print("Average stay without guy who stayed 12 years: \n", average_stay)
print("Average age: \n", average)
print("Percentage CoE: \n", p_CoE)
print("Outcome: ")
print("Percentage died: ", p_died)
print("Percentage recovered: ", p_recovered)
print("Occupations: \n", set(occupation_M))
print("Came from: ")
print("Percentage workhouse: ", p_workhouse)
print("Percentage police: ", p_police)

