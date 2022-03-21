import openpyxl

# load excel with its path
file_name = 'admissions_1901.xlsx'
wrkbk = openpyxl.load_workbook(file_name)

# get worksheet
sh = wrkbk.active

# calculate percentage 
def percentage(part, whole):
    return 100 * float(part)/float(whole)

male = 'Male'
female = 'Female'

# Create 2 arrays of pairs ('year', count), counting 
# admissions women and men in a given year
pair_women = []
pair_men = []

# Initialise arrays with count starting at 0 and year range 1860-1874
for i in range(1860, 1874):
    pair_women.append((str(i), 0))
    pair_men.append((str(i), 0))
pair_women = dict(pair_women)
pair_men = dict(pair_men)

# Count admissions of men and women in given years
for i in range(5, sh.max_row+1):
    for j in range(2, 3):
        first_col = sh.cell(row=i, column=1)
        if(first_col.value != None):
            cell_obj = sh.cell(row=i, column=j)
            for k in range(1860, 1874):
                if(cell_obj.value != None and str(k) in cell_obj.value):
                    if(cell_obj.value != None and female in cell_obj.value):
                        pair_women[str(k)] = pair_women[str(k)] + 1
                    elif(cell_obj.value != None and male in cell_obj.value):
                        pair_men[str(k)] = pair_men[str(k)] + 1

# Print data
total_year = 0
print("Patients admitted in 1860-1873 \n")
for i in range(1860, 1874):
    print("Number of women admitted in the year ", i)
    print(pair_women[str(i)])
    print("Number of men admitted in the year ", i)
    print(pair_men[str(i)])

    total_year = pair_women[str(i)] + pair_men[str(i)]
    p_women = percentage(pair_women[str(i)], total_year)
    p_men = percentage(pair_men[str(i)], total_year)
    print("Women: ", p_women, "%    Men: ", p_men, "%")
    print("\n")
