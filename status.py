import openpyxl

# load excel with its path
file_name = 'admissions_1901.xlsx'
wrkbk = openpyxl.load_workbook(file_name)

# get worksheet
sh = wrkbk.active

# calculate percentage 
def percentage(part, whole):
    return 100 * float(part)/float(whole)

male = 'Male'
female = 'Female'
single = 'Single'
ages = 'Age'
age = ''

# Create 2 arrays of pairs ('year', count, age, married, single, widow), counting 
# admissions women and men in a given year
pair_women = []
pair_men = []

year_age = []
year_single = []

# Initialise arrays with count starting at 0 and year range 1860-1874
for i in range(1852, 1900):
    pair_women.append((str(i), 0))
    pair_men.append((str(i), 0))
    year_age.append((str(i), 0))
    year_single.append((str(i), 0))
pair_women = dict(pair_women)
year_age = dict(year_age)
year_single = dict(year_single)
pair_men = dict(pair_men)

# Count admissions of men and women in given years
for i in range(5, sh.max_row+1):
    for j in range(2, 3):
        first_col = sh.cell(row=i, column=1)
        if(first_col.value != None):
            cell_obj = sh.cell(row=i, column=j)
            cell_obj2 = sh.cell(row=i+1, column=j)
            for k in range(1852, 1900):
                if(cell_obj.value != None and str(k) in cell_obj.value):
                    if(cell_obj.value != None and female in cell_obj.value):
                        pair_women[str(k)] = pair_women[str(k)] + 1

                        if(cell_obj2.value != None and single in cell_obj2.value):
                            year_single[str(k)] = year_single[str(k)] + 1
                        
                        """
                        if(cell_obj2.value != None and ages in cell_obj2.value):
                            for s in cell_obj2.value:
                                if s.isdigit():
                                    age = age + s
                                    print(cell_obj2.value, '\n')
                                if age != '':
                                    print(age, '\n')
                                    year_age[str(k)] = year_age[str(k)] + int(age)
                        """ 
                                

                    elif(cell_obj.value != None and male in cell_obj.value):
                        pair_men[str(k)] = pair_men[str(k)] + 1

            

# Print data
total_year = 0
lowest_p = 1000
highest_p = 0
print("Patients admitted in 1860-1873 \n")
for i in range(1852, 1900):
    print("Number of women admitted in the year ", i)
    print(pair_women[str(i)])
    print("Number of women single in the year and %", i)
    p_single = percentage(year_single[str(i)] , pair_women[str(i)] )
    if p_single < lowest_p:
        lowest_p = p_single
    if highest_p < p_single:
        highest_p = p_single
    print(year_single[str(i)], "     ", p_single, "%")
    print("Number of men admitted in the year ", i)
    print(pair_men[str(i)])

    total_year = pair_women[str(i)] + pair_men[str(i)]
    p_women = percentage(pair_women[str(i)], total_year)
    p_men = percentage(pair_men[str(i)], total_year)
    print("Women: ", p_women, "%    Men: ", p_men, "%")
    print("\n")

print ("lowest % : ", lowest_p, "       highest % : ", highest_p)
